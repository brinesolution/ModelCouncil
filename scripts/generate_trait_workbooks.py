from __future__ import annotations

import hashlib
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

PROJECT_ROOT = Path(__file__).resolve().parents[1]
TRAIT_ROOT = PROJECT_ROOT / "data" / "traits"
SOURCE_PATH = TRAIT_ROOT / "catalog.source.json"
MANIFEST_PATH = TRAIT_ROOT / "catalog.manifest.json"

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(sheet, columns: int) -> None:
    for cell in sheet[1][:columns]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _write_workbook(category: str, payload: dict) -> Path:
    rows = payload["rows"]
    if len(rows) != 10:
        raise ValueError(
            f"Trait category {category!r} must contain exactly 10 rows; got {len(rows)}"
        )

    columns = list(rows[0])
    if any(set(row) != set(columns) for row in rows):
        raise ValueError(f"Trait category {category!r} has inconsistent columns")

    workbook = Workbook()
    traits = workbook.active
    traits.title = "Traits"
    traits.append(columns)
    for row in rows:
        traits.append([row.get(column) for column in columns])

    _style_header(traits, len(columns))
    traits.freeze_panes = "A2"
    traits.auto_filter.ref = traits.dimensions

    for column_cells in traits.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        traits.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 40)
        for cell in column_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    metadata = workbook.create_sheet("Metadata")
    metadata_rows = [
        ("field", "value"),
        ("schema_version", "1.0"),
        ("category", category),
        ("workbook", f"{category}.xlsx"),
        ("purpose", payload.get("purpose", "")),
        ("weight_rule", "Enabled rows with weight > 0 are normalized within the category."),
        ("numeric_rule", "Normalized behavioural traits use 0.0-1.0 unless documented otherwise."),
        ("editable", "Preserve stable key values when editing labels, descriptions, weights, or parameters."),
    ]
    for row in metadata_rows:
        metadata.append(row)
    _style_header(metadata, 2)
    metadata.column_dimensions["A"].width = 24
    metadata.column_dimensions["B"].width = 64
    for row in metadata.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = TRAIT_ROOT / f"{category}.xlsx"
    workbook.save(output)
    return output


def generate() -> dict:
    source = json.loads(SOURCE_PATH.read_text(encoding="utf-8"))
    categories = source.get("categories", {})
    if not categories:
        raise ValueError("Trait source has no categories")

    manifest = {
        "schema_version": source.get("schema_version", "1.0"),
        "generated_on": "2026-08-17",
        "generator": "scripts/generate_trait_workbooks.py",
        "workbooks": [],
    }

    for category in sorted(categories):
        output = _write_workbook(category, categories[category])
        payload = output.read_bytes()
        manifest["workbooks"].append(
            {
                "file": output.name,
                "sha256": hashlib.sha256(payload).hexdigest(),
                "size_bytes": len(payload),
            }
        )

    MANIFEST_PATH.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
    return manifest


if __name__ == "__main__":
    result = generate()
    print(f"Generated {len(result['workbooks'])} trait workbooks in {TRAIT_ROOT}")
